import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# Load master dataset
df = pd.read_csv("dataset_master_07.08.csv")

# Get column indexes dynamically
cols = list(df.columns)
idx_order_no = cols.index("Order-No.")
idx_planned_step = cols.index("Planed-Master-Order-Processing-Position-No. as an ID")
idx_as_is_step = cols.index("As-Is-Master-Order-Processing-Position-No. as an ID")
idx_yield_qty = cols.index("Final Yield Quantity")
idx_scrap_qty = cols.index("Total Scrap Quantity")
idx_time_dev = None
if "Time_Deviation_Minutes" in cols:
    idx_time_dev = cols.index("Time_Deviation_Minutes")

# Create workbook & sheet
wb = Workbook()
ws = wb.active
ws.title = "Master Dataset"

# Add header
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    ws.append(row)
    if r_idx == 1:  # Header styling
        for cell in ws[r_idx]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F81BD")

# Apply color coding
for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
    order_no = row[idx_order_no].value
    yield_qty = row[idx_yield_qty].value
    scrap_qty = row[idx_scrap_qty].value
    planned_step = row[idx_planned_step].value
    as_is_step = row[idx_as_is_step].value
    time_dev = None
    if idx_time_dev is not None:
        try:
            time_dev = float(row[idx_time_dev].value)
        except:
            pass

    # Mode-based color for Order-No.
    if order_no.startswith("NEAT"):
        row[idx_order_no].fill = PatternFill("solid", fgColor="C6EFCE")
    elif order_no.startswith("MIXED"):
        row[idx_order_no].fill = PatternFill("solid", fgColor="FFEB9C")
    elif order_no.startswith("MISSING"):
        row[idx_order_no].fill = PatternFill("solid", fgColor="FFC7CE")
    elif order_no.startswith("OUT_OF_ORDER"):
        row[idx_order_no].fill = PatternFill("solid", fgColor="BDD7EE")
    elif order_no.startswith("EXTRA"):
        row[idx_order_no].fill = PatternFill("solid", fgColor="F4B084")
    elif order_no.startswith("DUPLICATES"):
        row[idx_order_no].fill = PatternFill("solid", fgColor="FFD966")
    elif order_no.startswith("DELAYED"):
        row[idx_order_no].fill = PatternFill("solid", fgColor="B4C6E7")
    elif order_no.startswith("QUANTITY"):
        row[idx_order_no].fill = PatternFill("solid", fgColor="E2EFDA")
    elif order_no.startswith("COMPLEX"):
        row[idx_order_no].fill = PatternFill("solid", fgColor="D9D2E9")

    # Yield colors
    if yield_qty is not None:
        if yield_qty >= 22:
            row[idx_yield_qty].fill = PatternFill("solid", fgColor="C6EFCE")
        elif yield_qty < 20:
            row[idx_yield_qty].fill = PatternFill("solid", fgColor="FFC7CE")

    # Scrap red if >0
    if scrap_qty and scrap_qty > 0:
        row[idx_scrap_qty].fill = PatternFill("solid", fgColor="FFC7CE")

    # Highlight mismatched steps
    if planned_step != as_is_step:
        row[idx_planned_step].fill = PatternFill("solid", fgColor="FFF2CC")
        row[idx_as_is_step].fill = PatternFill("solid", fgColor="FFF2CC")

    # Highlight time deviation
    if time_dev and time_dev > 0:
        row[idx_time_dev].fill = PatternFill("solid", fgColor="FCE4D6")

# Save Excel
wb.save("dataset_master_07.08.xlsx")
print("✅ Saved color-coded Excel → dataset_master_07.08.xlsx")
